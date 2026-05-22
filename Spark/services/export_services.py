
import sqlite3
import json
import os
import csv
from datetime import datetime

DB_NAME = "library.db"

# Переводчик месяцев для выгрузки заголовков отчетов
RU_MONTHS = {
    "01": "Январь", "02": "Февраль", "03": "Март", "04": "Апрель",
    "05": "Май", "06": "Июнь", "07": "Июль", "08": "Август",
    "09": "Сентябрь", "10": "Октябрь", "11": "Ноябрь", "12": "Декабрь"
}

def get_all_data():
    """Извлекает все таблицы из базы данных для формирования единого отчета"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # 1. Сбор Книг
    cursor.execute("SELECT inv_number, title, author, genre, status, place FROM books")
    books = cursor.fetchall()
    
    # 2. Сбор Читателей (с проверкой схемы на новые колонки)
    cursor.execute("PRAGMA table_info(readers)")
    columns = [col[1] for col in cursor.fetchall()]
    
    if "inv_card" in columns:
        cursor.execute("SELECT inv_card, full_name, reg_date, books_current, books_read FROM readers")
    else:
        cursor.execute("SELECT id, full_name, 'Не указана', books_current, books_read FROM readers")
    readers = cursor.fetchall()
    
    # 3. Сбор Истории (Транзакций)
    cursor.execute("SELECT inv_number, book_title, reader_name, action_type, timestamp FROM transactions ORDER BY id DESC")
    transactions = cursor.fetchall()
    
    conn.close()
    return books, readers, transactions


def export_to_json(target_dir="exports"):
    """Экспорт данных в структурированный JSON-формат"""
    os.makedirs(target_dir, exist_ok=True)
    books, readers, transactions = get_all_data()
    
    data = {
        "export_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "books": [{"inv_number": b[0], "title": b[1], "author": b[2], "genre": b[3], "status": b[4], "place": b[5]} for b in books],
        "readers": [{"library_card": r[0], "full_name": r[1], "registration_date": r[2], "books_on_hand": r[3], "total_read": r[4]} for r in readers],
        "history": [{"inv_number": t[0], "book_title": t[1], "reader_name": t[2], "action_type": t[3], "timestamp": t[4]} for t in transactions]
    }
    
    filepath = os.path.join(target_dir, f"library_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    return filepath


def export_to_excel(target_dir="exports"):
    """Экспорт в таблицы Excel (.xlsx). Если openpyxl не установлен, делает бэкап в .csv"""
    os.makedirs(target_dir, exist_ok=True)
    books, readers, transactions = get_all_data()
    
    filepath = os.path.join(target_dir, f"library_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # Фирменные стили из нашего Figma дизайна
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="294730", end_color="294730", fill_type="solid") # Наш темно-зеленый цвет
        data_font = Font(name="Arial", size=10)
        grid_side = Side(border_style="thin", color="D3D3D3")
        grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
        
        # --- Лист 1: Книжный Фонд ---
        ws1 = wb.active
        ws1.title = "Книжный фонд"
        ws1.append(["Инвентарный №", "Название книги", "Автор", "Жанр", "Статус", "Место хранения"])
        for row in books:
            ws1.append(row)
            
        # --- Лист 2: Читатели ---
        ws2 = wb.create_sheet(title="Читатели")
        ws2.append(["№ Билета / ID", "Полное ФИО", "Дата регистрации", "Книг на руках", "Всего прочитано"])
        for row in readers:
            ws2.append(row)
            
        # --- Лист 3: История операций ---
        ws3 = wb.create_sheet(title="История")
        ws3.append(["Инвентарный №", "Название книги", "ФИО Читателя", "Действие", "Дата и время"])
        for row in transactions:
            ws3.append(row)
            
        # Форматирование листов под дизайн приложения
        for ws in [ws1, ws2, ws3]:
            ws.views.sheetView[0].showGridLines = True
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = data_font
                    cell.border = grid_border
                    
            # Автоматическое расширение колонок по тексту
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(filepath)
        return filepath

    except ImportError:
        # Безопасный фоллбэк: если openpyxl не установлен, выгружаем в CSV
        csv_filepath = filepath.replace(".xlsx", ".csv")
        with open(csv_filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["--- КНИЖНЫЙ ФОНД ---"])
            writer.writerow(["Инвентарный №", "Название", "Автор", "Жанр", "Статус", "Место"])
            writer.writerows(books)
            writer.writerow([])
            writer.writerow(["--- ЧИТАТЕЛИ ---"])
            writer.writerow(["№ Билета", "ФИО", "Дата Регистрации", "На руках", "Прочитано"])
            writer.writerows(readers)
        return csv_filepath


def export_to_pdf(target_dir="exports"):
    """Экспорт в PDF через WeasyPrint. Если его нет, оставляет аккуратный HTML-отчет"""
    os.makedirs(target_dir, exist_ok=True)
    books, readers, transactions = get_all_data()
    
    filepath = os.path.join(target_dir, f"library_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    html_filepath = filepath.replace(".pdf", ".html")
    
    # Генерируем адаптивный HTML-шаблон с CSS-стилями нашего приложения
    html_content = f"""<!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
        @page {{ size: A4 landscape; margin: 12mm; }}
        body {{ font-family: 'Liberation Sans', 'Arial', sans-serif; color: #333; margin: 0; padding: 0; }}
        h1 {{ color: #294730; font-size: 24px; border-bottom: 3px solid #BEAC64; padding-bottom: 6px; margin-bottom: 20px; }}
        h2 {{ color: #294730; font-size: 16px; margin-top: 25px; margin-bottom: 8px; font-weight: bold; }}
        table {{ width: 100%; border-collapse: collapse; margin-bottom: 25px; font-size: 11px; page-break-inside: avoid; }}
        th {{ background-color: #294730; color: white; padding: 8px; text-align: left; font-weight: bold; border: 1px solid #294730; }}
        td {{ border: 1px solid #D3D3D3; padding: 6px; text-align: left; }}
        tr:nth-child(even) {{ background-color: #F5F4F2; }}
        .meta {{ font-size: 12px; color: #555; margin-bottom: 20px; }}
        .page-break {{ page-break-before: always; }}
    </style>
    </head>
    <body>
        <h1>Сводный отчет картотеки библиотекаря</h1>
        <div class="meta">Дата выгрузки: {datetime.now().strftime('%d.%m.%Y %H:%M')}</div>
        
        <h2>1. Книжный фонд фонда</h2>
        <table>
            <tr><th>Инвентарный №</th><th>Название книги</th><th>Автор</th><th>Жанр</th><th>Статус</th><th>Место</th></tr>
            {"".join(f"<tr><td>{b[0]}</td><td>{b[1]}</td><td>{b[2]}</td><td>{b[3]}</td><td>{b[4]}</td><td>{b[5]}</td></tr>" for b in books)}
        </table>
        
        <div class="page-break"></div>
        
        <h2>2. Список зарегистрированных читателей</h2>
        <table>
            <tr><th>№ Билета / ID</th><th>Полное ФИО</th><th>Дата регистрации</th><th>Книг на руках</th><th>Всего прочитано</th></tr>
            {"".join(f"<tr><td>{r[0]}</td><td>{r[1]}</td><td>{r[2]}</td><td>{r[3]}</td><td>{r[4]}</td></tr>" for r in readers)}
        </table>
        
        <div class="page-break"></div>
        
        <h2>3. История операций (Лог транзакций)</h2>
        <table>
            <tr><th>Инвентарный №</th><th>Название книги</th><th>ФИО Читателя</th><th>Действие</th><th>Дата и время</th></tr>
            {"".join(f"<tr><td>{t[0]}</td><td>{t[1]}</td><td>{t[2]}</td><td>{t[3]}</td><td>{t[4]}</td></tr>" for t in transactions)}
        </table>
    </body>
    </html>
    """
    
    with open(html_filepath, "w", encoding="utf-8") as f:
        f.write(html_content)
        
    try:
        from weasyprint import HTML
        HTML(html_filepath).write_pdf(filepath)
        # Если PDF успешно скомпилирован, удаляем промежуточный HTML-файл
        if os.path.exists(html_filepath):
            os.remove(html_filepath)
        return filepath
    except ImportError:
        # Если WeasyPrint не установлен, библиотекарь получит готовый интерактивный HTML-отчет
        return html_filepath
